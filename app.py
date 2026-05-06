import os
import re
import json
import imaplib
import email
import datetime
import traceback
import pandas as pd

from email.header import decode_header
from flask import Flask, jsonify, render_template_string, send_from_directory, request

import gspread
from google.oauth2.service_account import Credentials

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


app = Flask(__name__)

GMAIL_EMAIL = os.environ.get("GMAIL_EMAIL", "")
GMAIL_APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD", "")
GOOGLE_CREDS_JSON = os.environ.get("GOOGLE_CREDS_JSON", "")
GOOGLE_SHEET_NAME = os.environ.get("GOOGLE_SHEET_NAME", "mail_tracker")

CURRENT_YEAR = datetime.datetime.now().year
BASE_DIR = os.getcwd()

DOWNLOAD_PATH = os.path.join(BASE_DIR, "QTTY_RECAPS")
EDIT_PATH = os.path.join(DOWNLOAD_PATH, "Edit")

os.makedirs(DOWNLOAD_PATH, exist_ok=True)
os.makedirs(EDIT_PATH, exist_ok=True)


BLACK_BORDER = Border(
    top=Side(border_style="medium", color="000000"),
    bottom=Side(border_style="thin", color="000000"),
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000")
)


HOME_HTML = """
<!DOCTYPE html>
<html lang="ar" dir="rtl">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>QTTY Recap</title>
<style>
*{box-sizing:border-box}
body{
    margin:0;
    min-height:100vh;
    font-family:Arial,Tahoma,sans-serif;
    background:linear-gradient(135deg,#0f172a,#1e3a8a);
    display:flex;
    justify-content:center;
    align-items:center;
    color:white;
}
.card{
    width:94%;
    max-width:620px;
    background:rgba(255,255,255,0.12);
    border:1px solid rgba(255,255,255,0.2);
    border-radius:28px;
    padding:28px;
    text-align:center;
    box-shadow:0 20px 60px rgba(0,0,0,0.35);
}
h1{font-size:28px;margin-bottom:10px}
p{opacity:.9;line-height:1.7}
button{
    border:0;
    border-radius:14px;
    color:white;
    font-size:17px;
    font-weight:bold;
    cursor:pointer;
}
.run-btn{
    width:100%;
    padding:18px;
    background:#22c55e;
    font-size:22px;
    margin-top:18px;
}
#result{
    margin-top:20px;
    background:rgba(0,0,0,0.25);
    border-radius:16px;
    padding:15px;
    min-height:50px;
    white-space:pre-line;
    text-align:center;
}
.file-card{
    margin-top:14px;
    background:rgba(255,255,255,.12);
    border:1px solid rgba(255,255,255,.18);
    border-radius:16px;
    padding:14px;
    text-align:right;
}
.file-name{
    font-size:14px;
    opacity:.95;
    word-break:break-all;
    margin-bottom:12px;
}
.actions{
    display:flex;
    gap:8px;
    justify-content:center;
    flex-wrap:wrap;
}
.download-btn,.edit-btn,.save-btn,.cancel-btn{
    display:inline-flex;
    align-items:center;
    justify-content:center;
    gap:6px;
    padding:11px 14px;
    border-radius:12px;
    text-decoration:none;
    font-weight:bold;
    color:white;
    min-width:92px;
}
.download-btn{background:#38bdf8}
.edit-btn{background:#f59e0b}
.save-btn{background:#22c55e}
.cancel-btn{background:#64748b}
.edit-box{
    margin-top:12px;
    display:none;
    background:rgba(0,0,0,.22);
    border-radius:14px;
    padding:12px;
}
.inputs{
    display:flex;
    gap:8px;
    margin-bottom:10px;
}
input{
    width:100%;
    padding:12px;
    border-radius:10px;
    border:1px solid rgba(255,255,255,.25);
    background:rgba(255,255,255,.12);
    color:white;
    font-size:16px;
    text-align:center;
}
input::placeholder{color:rgba(255,255,255,.65)}
.small-note{
    font-size:13px;
    opacity:.8;
    margin-top:8px;
}
</style>
</head>
<body>
<div class="card">
    <h1>QTTY Recap</h1>
    <p>يقرأ الإيميلات الجديدة، يجهز ملفات الإكسيل، ويرقم الصفحات تلقائيًا حسب سنة الشحن من عمود Delivery Date.</p>
    <button class="run-btn" onclick="runProcess()">تشغيل الآن</button>
    <div id="result">جاهز للتشغيل ✅</div>
    <div id="files"></div>
</div>

<script>
function escapeHtml(text){
    return String(text)
        .replaceAll("&","&amp;")
        .replaceAll("<","&lt;")
        .replaceAll(">","&gt;")
        .replaceAll('"',"&quot;")
        .replaceAll("'","&#039;");
}

function showEditBox(id){
    document.getElementById("edit-box-" + id).style.display = "block";
}

function hideEditBox(id){
    document.getElementById("edit-box-" + id).style.display = "none";
}

async function savePageNumber(id, filename){
    const page = document.getElementById("page-" + id).value.trim();
    const year = document.getElementById("year-" + id).value.trim();

    if(!page || !year){
        alert("اكتب رقم الصفحة والسنة");
        return;
    }

    const res = await fetch("/update-page", {
        method:"POST",
        headers:{"Content-Type":"application/json"},
        body:JSON.stringify({filename, page, year})
    });

    const data = await res.json();

    if(!data.success){
        alert(data.message || "حدث خطأ أثناء تعديل رقم الصفحة");
        return;
    }

    const card = document.getElementById("file-card-" + id);
    card.querySelector(".file-name").innerText = data.filename;
    card.querySelector(".download-btn").href = data.download_url;

    hideEditBox(id);
    alert("تم تعديل رقم الصفحة داخل ملف الإكسيل ✅");
}

function renderFiles(files){
    const box = document.getElementById("files");
    box.innerHTML = "";

    files.forEach((file, index) => {
        const id = "f" + index;
        const safeName = escapeHtml(file.name);
        const page = escapeHtml(file.page || "");
        const year = escapeHtml(file.year || "");

        box.innerHTML += `
            <div class="file-card" id="file-card-${id}">
                <div class="file-name">${safeName}</div>
                <div class="actions">
                    <a class="download-btn" href="${file.download_url}">⬇️ تحميل</a>
                    <button class="edit-btn" onclick="showEditBox('${id}')">✏️ تعديل</button>
                </div>
                <div class="edit-box" id="edit-box-${id}">
                    <div class="inputs">
                        <input id="page-${id}" type="number" min="1" placeholder="رقم الصفحة" value="${page}">
                        <input id="year-${id}" type="number" min="2000" placeholder="السنة" value="${year}">
                    </div>
                    <div class="actions">
                        <button class="save-btn" onclick="savePageNumber('${id}', '${safeName}')">حفظ</button>
                        <button class="cancel-btn" onclick="hideEditBox('${id}')">إلغاء</button>
                    </div>
                    <div class="small-note">التعديل يغيّر عنوان الصفحة داخل هذا الملف فقط ولا يغيّر عداد Google Sheet.</div>
                </div>
            </div>
        `;
    });
}

async function runProcess(){
    const result = document.getElementById("result");
    const filesBox = document.getElementById("files");
    result.innerText = "جاري التشغيل... انتظر";
    filesBox.innerHTML = "";

    try{
        const res = await fetch("/run");
        const data = await res.json();

        result.innerHTML = data.message.replaceAll("\\n", "<br>");

        if(data.success && data.files && data.files.length){
            renderFiles(data.files);
        }

    }catch(e){
        result.innerText = "حدث خطأ أثناء التشغيل: " + e;
    }
}
</script>
</body>
</html>
"""


def log(msg):
    print(f"[{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}", flush=True)


def decode_mime_text(value):
    if not value:
        return ""

    decoded_parts = decode_header(value)
    result = []

    for part, encoding in decoded_parts:
        if isinstance(part, bytes):
            try:
                result.append(part.decode(encoding or "utf-8", errors="replace"))
            except Exception:
                result.append(part.decode("latin-1", errors="replace"))
        else:
            result.append(str(part))

    return "".join(result)


def clean_filename(filename: str) -> str:
    if isinstance(filename, bytes):
        try:
            filename = filename.decode("utf-8")
        except UnicodeDecodeError:
            filename = filename.decode("latin-1", errors="replace")

    return re.sub(r'[\\/*?:"<>|]', "_", str(filename)).strip()


def safe_download_filename(filename: str) -> str:
    filename = os.path.basename(str(filename))
    filename = clean_filename(filename)
    if not filename.lower().endswith(".xlsx"):
        raise Exception("اسم الملف غير صالح")
    return filename


def get_google_sheet():
    if not GOOGLE_CREDS_JSON:
        raise Exception("متغير GOOGLE_CREDS_JSON غير موجود في Render Environment Variables")

    creds_dict = json.loads(GOOGLE_CREDS_JSON)

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]

    creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
    client = gspread.authorize(creds)

    return client.open(GOOGLE_SHEET_NAME).sheet1


def ensure_sheet_structure(sheet):
    sheet.update("A1:B1", [["year", "page"]])
    sheet.update("D1:H1", [["message_id", "file_name", "year", "page", "created_at"]])


def get_year_rows(sheet):
    ensure_sheet_structure(sheet)

    years = sheet.col_values(1)
    pages = sheet.col_values(2)

    clean_rows = []
    max_len = max(len(years), len(pages))

    for i in range(2, max_len + 1):
        year = ""
        page = ""

        if i - 1 < len(years):
            year = str(years[i - 1]).strip()

        if i - 1 < len(pages):
            page = str(pages[i - 1]).strip()

        if year.isdigit():
            try:
                page_num = int(float(page)) if page else 0
            except Exception:
                page_num = 0

            clean_rows.append({
                "row_index": i,
                "year": year,
                "page": page_num
            })

    return clean_rows


def get_next_page_for_year(file_year: int) -> dict:
    sheet = get_google_sheet()
    ensure_sheet_structure(sheet)

    year_str = str(file_year)
    suffix = year_str[-2:]

    rows = get_year_rows(sheet)

    target = None
    for row in rows:
        if row["year"] == year_str:
            target = row
            break

    if target is None:
        next_empty_row = 2

        used_rows = sheet.col_values(1)
        if len(used_rows) >= 2:
            next_empty_row = len(used_rows) + 1

        sheet.update(f"A{next_empty_row}:B{next_empty_row}", [[year_str, 0]])
        current_page = 0
        row_index = next_empty_row
    else:
        current_page = target["page"]
        row_index = target["row_index"]

    next_page = current_page + 1

    sheet.update(f"B{row_index}", [[next_page]])

    return {
        "page": next_page,
        "year": year_str,
        "suffix": suffix,
        "page_title": f"PAGE {next_page}/{suffix}",
        "file_name": f"PAGE {next_page}-{suffix}"
    }


def get_processed_message_ids() -> set:
    sheet = get_google_sheet()
    ensure_sheet_structure(sheet)

    values = sheet.col_values(4)
    return set(str(v).strip() for v in values[1:] if str(v).strip())


def get_next_log_row(sheet):
    message_ids = sheet.col_values(4)
    return len(message_ids) + 1 if len(message_ids) >= 1 else 2


def append_processed_log(message_id: str, file_name: str, year: str, page: str):
    sheet = get_google_sheet()
    ensure_sheet_structure(sheet)

    created_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row = get_next_log_row(sheet)

    sheet.update(
        f"D{row}:H{row}",
        [[message_id, file_name, year, page, created_at]]
    )


def normalize_header(value):
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def extract_year_from_delivery_date_value(value):
    if value is None:
        return None

    if isinstance(value, (datetime.datetime, datetime.date)):
        return int(value.year)

    text = str(value).strip()

    if not text:
        return None

    parsed = pd.to_datetime(text, errors="coerce", dayfirst=True)

    if pd.notna(parsed):
        return int(parsed.year)

    year_match = re.search(r"(20\d{2})", text)
    if year_match:
        return int(year_match.group(1))

    short_year_match = re.search(r"(?:/|-|\.)(\d{2})$", text)
    if short_year_match:
        return int("20" + short_year_match.group(1))

    return None


def extract_file_year_from_excel(filepath: str, fallback_year: int) -> int:
    file_year = fallback_year

    try:
        wb_temp = load_workbook(filepath, read_only=True, data_only=True)
        ws_temp = wb_temp.active

        target_col_idx = None

        for cell in ws_temp[1]:
            header = normalize_header(cell.value)

            if header in ["delivery date", "deliverydate"]:
                target_col_idx = cell.column
                break

            if "delivery" in header and "date" in header:
                target_col_idx = cell.column
                break

        if target_col_idx:
            for row_num in range(2, min(ws_temp.max_row, 20) + 1):
                cell_val = ws_temp.cell(row=row_num, column=target_col_idx).value
                extracted_year = extract_year_from_delivery_date_value(cell_val)

                if extracted_year:
                    file_year = extracted_year
                    break

        wb_temp.close()

    except Exception as e:
        log(f"Year extraction warning: {e}")

    return int(file_year)


def get_message_real_id(msg, fallback_id: str) -> str:
    real_id = msg.get("Message-ID", "")
    if real_id:
        return real_id.strip()
    return str(fallback_id)


def download_attachments() -> list:
    if not GMAIL_EMAIL or not GMAIL_APP_PASSWORD:
        raise Exception("GMAIL_EMAIL أو GMAIL_APP_PASSWORD غير موجودين في Render Environment Variables")

    downloaded_files = []
    processed_ids = get_processed_message_ids()

    log("Connecting to Gmail...")

    mail = imaplib.IMAP4_SSL("imap.gmail.com")
    mail.login(GMAIL_EMAIL, GMAIL_APP_PASSWORD)
    mail.select("INBOX")

    status, messages = mail.search(None, 'UNSEEN SUBJECT "Qtty Recap"')

    if status != "OK" or not messages or not messages[0]:
        mail.logout()
        return []

    all_message_ids = messages[0].split()

    for num_bytes in all_message_ids:
        num_str = num_bytes.decode()

        try:
            status, msg_data = mail.fetch(num_bytes, "(RFC822)")

            if status != "OK" or not msg_data or not msg_data[0]:
                continue

            msg = email.message_from_bytes(msg_data[0][1])
            real_message_id = get_message_real_id(msg, num_str)

            if real_message_id in processed_ids:
                log(f"Skipped already processed email: {real_message_id}")
                mail.store(num_bytes, '+FLAGS', '\\Seen')
                continue

            subject = decode_mime_text(msg.get("Subject", ""))
            log(f"Processing email: {subject}")

            date_str = msg.get("Date", "")
            fallback_year = CURRENT_YEAR
            local_date_prefix = datetime.datetime.now().strftime("%Y%m%d")

            if date_str:
                try:
                    date_tuple = email.utils.parsedate_tz(date_str)
                    if date_tuple:
                        local_timestamp = email.utils.mktime_tz(date_tuple)
                        local_dt = datetime.datetime.fromtimestamp(local_timestamp)
                        fallback_year = local_dt.year
                        local_date_prefix = local_dt.strftime("%Y%m%d")
                except Exception:
                    pass

            files_saved_from_this_email = 0

            for part in msg.walk():
                if part.get_content_maintype() == "multipart":
                    continue

                if part.get("Content-Disposition") is None:
                    continue

                filename = part.get_filename()

                if not filename:
                    continue

                original_filename = decode_mime_text(filename)

                if not original_filename.lower().endswith((".xlsx", ".xls")):
                    continue

                clean_name = clean_filename(f"{local_date_prefix}_{original_filename}")
                filepath = os.path.join(DOWNLOAD_PATH, clean_name)

                payload = part.get_payload(decode=True)

                if not payload:
                    continue

                with open(filepath, "wb") as f:
                    f.write(payload)

                file_year = extract_file_year_from_excel(filepath, fallback_year)

                downloaded_files.append({
                    "path": filepath,
                    "file_name": clean_name,
                    "year": file_year,
                    "message_id": real_message_id
                })

                files_saved_from_this_email += 1
                log(f"Downloaded: {clean_name} | Delivery Year: {file_year}")

            if files_saved_from_this_email > 0:
                mail.store(num_bytes, '+FLAGS', '\\Seen')
                processed_ids.add(real_message_id)
                log(f"Marked email as read: {num_str}")

        except Exception as e:
            log(f"Error processing email {num_str}: {e}")
            log(traceback.format_exc())

    mail.close()
    mail.logout()

    return downloaded_files


def preserve_images(ws):
    return list(getattr(ws, "_images", []))


def restore_images(ws, images):
    try:
        ws._images = images
    except Exception as e:
        log(f"Image restore warning: {e}")


def set_sheet_page_title(excel_file_path: str, page: int, year: int):
    wb = load_workbook(excel_file_path)
    ws = wb.active

    suffix = str(year)[-2:]

    ws.cell(row=1, column=1).value = f"PAGE {page}/{suffix}"

    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(name="Arial", color="FF0000", bold=True, size=36)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(excel_file_path)


def transform_excel_file(input_file: str, page_title: str, output_file_name: str) -> str | None:
    try:
        output_file_path = os.path.join(EDIT_PATH, f"{output_file_name}_edit.xlsx")

        wb = load_workbook(input_file)
        ws = wb.active

        saved_images = preserve_images(ws)

        df = pd.read_excel(input_file, engine="openpyxl")

        header_row_index = 1
        headers = [cell.value for cell in ws[header_row_index] if cell.value is not None]

        keywords_to_delete = ["customer", "price", "factory"]
        cols_to_delete_indices = []

        for idx, header in enumerate(headers):
            if header and any(kw.lower() in str(header).lower() for kw in keywords_to_delete):
                cols_to_delete_indices.append(idx + 1)

        for col_idx in sorted(cols_to_delete_indices, reverse=True):
            ws.delete_cols(col_idx)

        current_num_columns = ws.max_column
        ws.insert_rows(1)

        title_cell = ws.cell(row=1, column=1)
        title_cell.value = page_title
        title_cell.font = Font(name="Arial", color="FF0000", bold=True, size=36)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=current_num_columns
        )

        ws["A2"].fill = PatternFill(fill_type=None)

        if "Style # " not in df.columns or "Quantity Ordered" not in df.columns:
            raise Exception("الأعمدة المطلوبة غير موجودة: Style # و Quantity Ordered")

        df["Style_number"] = df["Style # "].astype(str)
        df["last4"] = df["Style_number"].str.slice(-4)
        df["Quantity Ordered"] = pd.to_numeric(
            df["Quantity Ordered"],
            errors="coerce"
        ).fillna(0)

        grouped = df.groupby("last4")["Quantity Ordered"].sum().reset_index()

        insert_col_index = ws.max_column
        ws.insert_cols(insert_col_index + 1, 2)

        total_col_letter = get_column_letter(insert_col_index + 1)
        model_col_letter = get_column_letter(insert_col_index + 2)

        header_row_num = 2

        ws[f"{total_col_letter}{header_row_num}"].value = "الإجمالي"
        ws[f"{model_col_letter}{header_row_num}"].value = "الموديل"

        header_font = Font(color="FFFFFF", bold=True)
        header_fill = PatternFill(
            start_color="4a90e2",
            end_color="4a90e2",
            fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_letter in [total_col_letter, model_col_letter]:
            cell = ws[f"{col_letter}{header_row_num}"]
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = BLACK_BORDER

        cell_font = Font(name="Arial", size=20, color="FF0000", bold=True)
        cell_alignment = Alignment(horizontal="center", vertical="center")

        visited_last4 = set()
        base_row_offset = 3

        for last4_val in df["last4"]:
            if last4_val in visited_last4:
                continue

            group_df_indices = df[df["last4"] == last4_val].index.tolist()

            if not group_df_indices:
                continue

            start_excel_row = group_df_indices[0] + base_row_offset
            end_excel_row = group_df_indices[-1] + base_row_offset

            total_qty_series = grouped[grouped["last4"] == last4_val]["Quantity Ordered"]
            total_qty = total_qty_series.values[0] if not total_qty_series.empty else 0

            total_cell = ws[f"{total_col_letter}{start_excel_row}"]
            total_cell.value = total_qty
            total_cell.font = cell_font
            total_cell.alignment = cell_alignment

            if end_excel_row >= start_excel_row:
                ws.merge_cells(
                    start_row=start_excel_row,
                    start_column=insert_col_index + 1,
                    end_row=end_excel_row,
                    end_column=insert_col_index + 1
                )

            for row_num in range(start_excel_row, end_excel_row + 1):
                ws.cell(row=row_num, column=insert_col_index + 1).border = BLACK_BORDER

            model_cell = ws[f"{model_col_letter}{start_excel_row}"]
            model_cell.value = last4_val
            model_cell.font = cell_font
            model_cell.alignment = cell_alignment

            if end_excel_row >= start_excel_row:
                ws.merge_cells(
                    start_row=start_excel_row,
                    start_column=insert_col_index + 2,
                    end_row=end_excel_row,
                    end_column=insert_col_index + 2
                )

            for row_num in range(start_excel_row, end_excel_row + 1):
                ws.cell(row=row_num, column=insert_col_index + 2).border = BLACK_BORDER

            visited_last4.add(last4_val)

        current_headers = [cell.value for cell in ws[2]]

        column_widths = {
            "po #": 20,
            "style #": 18,
            "size": 15,
            "date": 12,
            "breakdown": 12,
            "quantity ordered": 12,
            "poly bag": 10,
            "الإجمالي": 12,
            "الموديل": 12
        }

        for col_idx, header_val in enumerate(current_headers):
            if not header_val:
                continue

            col_letter = get_column_letter(col_idx + 1)
            header_lower = str(header_val).lower()

            for key, width in column_widths.items():
                if key.lower() in header_lower:
                    ws.column_dimensions[col_letter].width = width
                    break

        restore_images(ws, saved_images)

        wb.save(output_file_path)
        log(f"Saved transformed file: {output_file_path}")

        return output_file_path

    except Exception as e:
        log(f"Transform error: {e}")
        log(traceback.format_exc())
        return None


def process_all():
    downloaded_files = download_attachments()

    edit_files = []

    for item in downloaded_files:
        file_path = item["path"]
        file_year = int(item["year"])
        original_file_name = item["file_name"]
        message_id = item["message_id"]

        page_data = get_next_page_for_year(file_year)

        transformed_file = transform_excel_file(
            input_file=file_path,
            page_title=page_data["page_title"],
            output_file_name=page_data["file_name"]
        )

        if transformed_file:
            final_name = os.path.basename(transformed_file)

            append_processed_log(
                message_id=message_id,
                file_name=original_file_name,
                year=str(file_year),
                page=str(page_data["page"])
            )

            edit_files.append({
                "name": final_name,
                "page": page_data["page"],
                "year": file_year,
                "download_url": f"/download-file/{final_name}"
            })

    return {
        "downloaded": len(downloaded_files),
        "edit_files_count": len(edit_files),
        "edit_files": edit_files
    }


@app.route("/")
def home():
    return render_template_string(HOME_HTML)


@app.route("/run")
def run():
    try:
        result = process_all()

        if result["edit_files_count"] == 0:
            msg = (
                "تم التشغيل بنجاح ✅\\n"
                "لا توجد ملفات إكسيل جديدة جاهزة للتحميل الآن."
            )
        else:
            msg = (
                "تم التشغيل بنجاح ✅\\n"
                f"تم تحميل {result['downloaded']} ملف من الإيميل.\\n"
                f"تم تجهيز {result['edit_files_count']} ملف إكسيل.\\n\\n"
                "يمكنك تحميل كل ملف مباشرة أو تعديل رقم الصفحة قبل التحميل."
            )

        return jsonify({
            "success": True,
            "message": msg,
            "files": result["edit_files"]
        })

    except Exception as e:
        log(traceback.format_exc())

        return jsonify({
            "success": False,
            "message": f"حدث خطأ ❌\\n{str(e)}"
        }), 500


@app.route("/download-file/<filename>")
def download_file(filename):
    try:
        safe_name = safe_download_filename(filename)
        return send_from_directory(
            EDIT_PATH,
            safe_name,
            as_attachment=True
        )
    except Exception as e:
        return jsonify({
            "success": False,
            "message": str(e)
        }), 400


@app.route("/update-page", methods=["POST"])
def update_page():
    try:
        data = request.get_json(force=True)

        old_filename = safe_download_filename(data.get("filename", ""))
        page = int(data.get("page", 0))
        year = int(data.get("year", 0))

        if page <= 0:
            raise Exception("رقم الصفحة غير صالح")

        if year < 2000 or year > 2099:
            raise Exception("السنة غير صالحة")

        old_path = os.path.join(EDIT_PATH, old_filename)

        if not os.path.exists(old_path):
            raise Exception("الملف غير موجود")

        suffix = str(year)[-2:]
        new_filename = f"PAGE {page}-{suffix}_edit.xlsx"
        new_path = os.path.join(EDIT_PATH, new_filename)

        if old_path != new_path:
            os.replace(old_path, new_path)

        set_sheet_page_title(new_path, page, year)

        return jsonify({
            "success": True,
            "message": "تم تعديل رقم الصفحة داخل الملف بنجاح",
            "filename": new_filename,
            "page": page,
            "year": year,
            "download_url": f"/download-file/{new_filename}"
        })

    except Exception as e:
        log(traceback.format_exc())

        return jsonify({
            "success": False,
            "message": str(e)
        }), 400


@app.route("/health")
def health():
    return jsonify({"status": "ok"})


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
